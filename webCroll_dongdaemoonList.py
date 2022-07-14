from selenium import webdriver #selenium 드라이버 import
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
#key는 값을 넣는데 사용됨
#by는 얻어오는데 사용됨
import time
import openpyxl as xl

driver = webdriver.Chrome('./chromedriver')
driver.implicitly_wait(1) # 로딩 지연
driver.get('https://kiwikiwi.kr/stores?area=%EB%8F%99%EB%8C%80%EB%AC%B8%EC%A2%85%ED%95%A9%EC%8B%9C%EC%9E%A5')
driver.maximize_window()
time.sleep(1)

for i in range(1,6000):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    print(i)

store_name = driver.find_elements_by_class_name('store-name')
store_address = driver.find_elements_by_class_name('store-address')


#print(type(store_name))
wb = xl.Workbook()
ws = wb.active
ws.title = "result_"
ws["A1"] = "Name"
ws["B1"] = "Address"

count = 2
#for i in store_name:
print(len(store_name))
for i in store_name:
    ws['A' + str(count)] = i.text
    print(i.text)
    count += 1

count = 2
#for i in store_address:
for i in store_address:
    ws['B' + str(count)] = i.text
    print(i.text)
    count += 1
print("complete")
wb.save(r"\\192.168.0.100\data\TFT\200.saData\동대문매장리스트.xlsx")
print("complete_save")